import os
import pdfplumber
from openpyxl import Workbook
from pathlib import Path

def combine_pdfs_to_xlsx(input_dir, output_file, single_sheet=True):
    """
    Combine data from all PDF files in input_dir into a single XLSX file.
    Args:
        input_dir (str): Directory containing PDF files.
        output_file (str): Path to output XLSX file (e.g., 'project_data.xlsx').
        single_sheet (bool): If True, combine all data into one sheet; else, use separate sheets per PDF.
    """
    # Create workbook
    wb = Workbook()
    if single_sheet:
        ws = wb.active
        ws.title = "Project Data"
    else:
        wb.remove(wb.active)  # Remove default sheet

    # Ensure output directory exists
    output_dir = os.path.dirname(output_file)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    row = 1  # Track row for single-sheet mode
    pdf_files = sorted(Path(input_dir).glob("*.pdf"))  # Sort for consistency
    if not pdf_files:
        print(f"No PDF files found in {input_dir}")
        return

    for idx, pdf_file in enumerate(pdf_files, start=1):
        # Create sheet for multi-sheet mode
        if not single_sheet:
            ws = wb.create_sheet(title=f"PDF_{idx}")

        # Add header for each PDF in single-sheet mode
        if single_sheet:
            ws.cell(row=row, column=1, value=f"Data from {pdf_file.name}")
            row += 1

        # Open PDF
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                # Try extracting tables
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for table_row in table:
                            for col, cell in enumerate(table_row, start=1):
                                ws.cell(row=row, column=col, value=cell or "")
                            row += 1
                        row += 1  # Blank row between tables
                else:
                    # Fall back to text extraction
                    text = page.extract_text()
                    if text:
                        lines = text.split("\n")
                        for line in lines:
                            ws.cell(row=row, column=1, value=line.strip())
                            row += 1
                    row += 1  # Blank row between pages

        # Add blank row between PDFs in single-sheet mode
        if single_sheet:
            row += 1

    # Save XLSX
    wb.save(output_file)
    print(f"Combined {len(pdf_files)} PDFs into {output_file}")

if __name__ == "__main__":
    # Define input and output paths
    input_dir = "pdfs"  # Directory with PDF files
    output_file = "xlsx_output/project_data.xlsx"  # Output XLSX file

    # Run conversion (single sheet)
    combine_pdfs_to_xlsx(input_dir, output_file, single_sheet=True)

    # Optional: Run for separate sheets
    # combine_pdfs_to_xlsx(input_dir, output_file, single_sheet=False)